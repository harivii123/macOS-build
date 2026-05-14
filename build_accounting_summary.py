
#!/usr/bin/env python3

from __future__ import annotations

import argparse
import math
import re
from collections import defaultdict
from copy import copy
from datetime import datetime, date
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


MAIN_SHEET_NAME = "Daily Finances"
FILE_PATTERNS = ("Daily Finances *.xlsm", "Daily Finances *.xlsx")


def clean_text(value: Any) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        return value or None
    return str(value).strip() or None


def to_number(value: Any) -> float:
    if value is None:
        return 0.0
    if isinstance(value, bool):
        return float(value)
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        text = value.strip().replace(",", "")
        if not text:
            return 0.0
        try:
            return float(text)
        except ValueError:
            return 0.0
    return 0.0


def maybe_number(value: Any) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, str) and not value.strip():
        return None
    n = to_number(value)
    return n


def rounded(value: float) -> float | int:
    if abs(value - round(value)) < 1e-9:
        return int(round(value))
    return round(value, 2)


def parse_file_date(path: Path) -> date:
    match = re.search(r"(\d{4}-\d{2}-\d{2})", path.name)
    if not match:
        raise ValueError(f"Could not find YYYY-MM-DD date in filename: {path.name}")
    return datetime.strptime(match.group(1), "%Y-%m-%d").date()


def safe_sheet(workbook) :
    if MAIN_SHEET_NAME in workbook.sheetnames:
        return workbook[MAIN_SHEET_NAME]
    return workbook[workbook.sheetnames[0]]


def copy_row_style(source_ws, source_row: int, target_ws, target_row: int, max_col: int) -> None:
    for col in range(1, max_col + 1):
        source = source_ws.cell(source_row, col)
        target = target_ws.cell(target_row, col)
        if source.has_style:
            target._style = copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        if source.alignment:
            target.alignment = copy(source.alignment)
        if source.font:
            target.font = copy(source.font)
        if source.fill:
            target.fill = copy(source.fill)
        if source.border:
            target.border = copy(source.border)


def group_sum_count(rows: Iterable[dict], key_name: str, amount_name: str) -> List[Tuple[str, int, float]]:
    grouped = defaultdict(lambda: {"entries": 0, "amount": 0.0})
    for row in rows:
        key = clean_text(row.get(key_name)) or "(Blank)"
        grouped[key]["entries"] += 1
        grouped[key]["amount"] += to_number(row.get(amount_name))
    items = [(key, values["entries"], rounded(values["amount"])) for key, values in grouped.items()]
    items.sort(key=lambda x: (-to_number(x[2]), x[0].lower()))
    return items


def infer_money_to_bank(opening: Optional[float], income: Optional[float], expense: Optional[float], closing: Optional[float]) -> float:
    if None in (opening, income, expense, closing):
        return 0.0
    inferred = float(opening) + float(income) - float(expense) - float(closing)
    if abs(inferred) < 0.5:
        return 0.0
    return rounded(inferred)


def make_unique_period_labels(base_labels: List[str]) -> List[str]:
    seen = defaultdict(int)
    result = []
    for label in base_labels:
        seen[label] += 1
        if seen[label] == 1:
            result.append(label)
        else:
            result.append(f"{label} #{seen[label]}")
    return result


def extract_income_rows(ws, file_date: date) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    for r in range(2, ws.max_row + 1):
        raw = {col: ws[f"{col}{r}"].value for col in "ABCDEFGHIJK"}

        row = {
            "date": file_date,
            "name": clean_text(raw["B"]),
            "payment_status": clean_text(raw["D"]),
            "services": clean_text(raw["E"]),
            "payment_method": clean_text(raw["F"]),
            "diving_fees": maybe_number(raw["G"]),
            "extra_misc": maybe_number(raw["H"]),
            "deposit": maybe_number(raw["I"]),
            "discount": maybe_number(raw["J"]),
            "recorded_total": maybe_number(raw["K"]),
        }

        if row["recorded_total"] is None:
            calc_total = (
                to_number(raw["G"]) +
                to_number(raw["H"]) -
                to_number(raw["I"]) -
                to_number(raw["J"])
            )
            row["recorded_total"] = rounded(calc_total) if abs(calc_total) > 1e-9 else None

        meaningful = any(
            row[field] is not None
            for field in (
                "name", "payment_status", "services", "payment_method",
                "diving_fees", "extra_misc", "deposit", "discount", "recorded_total"
            )
        )

        if meaningful and row["recorded_total"] not in (None, 0):
            row["recorded_total"] = rounded(to_number(row["recorded_total"]))
            if row["diving_fees"] is not None:
                row["diving_fees"] = rounded(to_number(row["diving_fees"]))
            if row["extra_misc"] is not None:
                row["extra_misc"] = rounded(to_number(row["extra_misc"]))
            if row["deposit"] is not None:
                row["deposit"] = rounded(to_number(row["deposit"]))
            if row["discount"] is not None:
                row["discount"] = rounded(to_number(row["discount"]))
            rows.append(row)

    return rows


def extract_expense_rows(ws, file_date: date) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    for r in range(2, ws.max_row + 1):
        raw = {col: ws[f"{col}{r}"].value for col in "LMNO"}

        row = {
            "date": file_date,
            "expense_category": clean_text(raw["L"]),
            "detail": clean_text(raw["M"]),
            "amount": maybe_number(raw["N"]),
            "payment_method": clean_text(raw["O"]),
        }

        meaningful = any(
            row[field] is not None
            for field in ("expense_category", "detail", "amount", "payment_method")
        )

        if meaningful and row["amount"] not in (None, 0):
            row["amount"] = rounded(to_number(row["amount"]))
            rows.append(row)

    return rows


def extract_daily_cashbox(ws, file_date: date) -> Dict[str, Any]:
    cash_income = maybe_number(ws["Q4"].value) or 0
    cash_expense = maybe_number(ws["Q5"].value) or 0
    opening_balance = maybe_number(ws["Q8"].value) or 0
    closing_balance = maybe_number(ws["Q9"].value) or 0

    explicit_money_to_bank = maybe_number(ws["Q6"].value)
    if explicit_money_to_bank is None:
        money_to_bank = infer_money_to_bank(opening_balance, cash_income, cash_expense, closing_balance)
    else:
        money_to_bank = rounded(to_number(explicit_money_to_bank))

    net_before_bank = rounded(cash_income - cash_expense)
    net_after_bank = rounded(cash_income - cash_expense - money_to_bank)

    return {
        "date": file_date,
        "opening_balance": rounded(opening_balance),
        "cash_income": rounded(cash_income),
        "cash_expense": rounded(cash_expense),
        "money_to_bank": rounded(money_to_bank),
        "net_cash_before_bank": net_before_bank,
        "net_cash_after_bank": net_after_bank,
        "closing_balance": rounded(closing_balance),
    }


def assign_periods(daily_rows: List[Dict[str, Any]]) -> None:
    """
    Assign a period label to each daily row.
    A new period starts when the cashbox chain breaks:
    current opening balance != previous closing balance.
    """
    if not daily_rows:
        return

    daily_rows.sort(key=lambda row: row["date"])

    base_labels = []
    period_starts = []
    previous = None
    for row in daily_rows:
        start_new = False
        if previous is None:
            start_new = True
        else:
            prev_close = to_number(previous.get("closing_balance"))
            curr_open = to_number(row.get("opening_balance"))
            if abs(prev_close - curr_open) > 0.5:
                start_new = True

        if start_new:
            period_starts.append(len(base_labels))
            base_labels.append(row["date"].strftime("%b %Y block"))
        previous = row

    unique_labels = make_unique_period_labels(base_labels)

    period_index = -1
    previous = None
    for row in daily_rows:
        if previous is None:
            period_index += 1
        else:
            prev_close = to_number(previous.get("closing_balance"))
            curr_open = to_number(row.get("opening_balance"))
            if abs(prev_close - curr_open) > 0.5:
                period_index += 1
        row["period"] = unique_labels[period_index]
        previous = row


TITLE_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
SECTION_FILL = PatternFill("solid", fgColor="EDEDED")
WHITE_FONT = Font(color="FFFFFF", bold=True, size=12)
TITLE_FONT = Font(color="FFFFFF", bold=True, size=13)
HEADER_FONT = Font(bold=True)
BOLD_FONT = Font(bold=True)
THIN_GRAY = Side(style="thin", color="D9D9D9")
BOTTOM_BORDER = Border(bottom=Side(style="medium", color="808080"))


def style_title(ws, title: str, width: int) -> None:
    ws.cell(1, 1, title)
    ws.cell(1, 1).font = TITLE_FONT
    ws.cell(1, 1).fill = TITLE_FILL
    ws.cell(1, 1).alignment = Alignment(horizontal="left")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=width)


def style_header_row(ws, row_num: int, width: int) -> None:
    for col in range(1, width + 1):
        cell = ws.cell(row_num, col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(bottom=THIN_GRAY)


def autofit_columns(ws, min_width: int = 10, max_width: int = 45) -> None:
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        letter = get_column_letter(col_idx)
        for cell in ws[letter]:
            value = cell.value
            if value is None:
                continue
            if isinstance(value, (datetime, date)):
                text = value.strftime("%Y-%m-%d")
            else:
                text = str(value)
            max_len = max(max_len, len(text))
        ws.column_dimensions[letter].width = min(max(max_len + 2, min_width), max_width)


def apply_number_formats(ws, date_cols: Iterable[int] = (), currency_cols: Iterable[int] = (), percent_cols: Iterable[int] = ()) -> None:
    for col in date_cols:
        for row in range(3, ws.max_row + 1):
            ws.cell(row, col).number_format = "yyyy-mm-dd"
    for col in currency_cols:
        for row in range(3, ws.max_row + 1):
            ws.cell(row, col).number_format = '#,##0;[Red](#,##0);-'
    for col in percent_cols:
        for row in range(3, ws.max_row + 1):
            ws.cell(row, col).number_format = "0.0%"


def write_table_sheet(
    wb: Workbook,
    title: str,
    sheet_name: str,
    headers: List[str],
    rows: List[List[Any]],
    date_cols: Iterable[int] = (),
    currency_cols: Iterable[int] = (),
    percent_cols: Iterable[int] = (),
    freeze_panes: str = "A3",
) -> None:
    ws = wb.create_sheet(title=sheet_name)
    style_title(ws, title, len(headers))
    ws.append(headers)
    style_header_row(ws, 2, len(headers))

    for row in rows:
        ws.append(row)

    apply_number_formats(ws, date_cols=date_cols, currency_cols=currency_cols, percent_cols=percent_cols)
    ws.freeze_panes = freeze_panes
    ws.auto_filter.ref = ws.dimensions
    autofit_columns(ws)


def build_output_path(input_dir: Path, min_date: date, max_date: date, explicit_output: Optional[str]) -> Path:
    if explicit_output:
        return (input_dir / explicit_output).resolve()

    if min_date.year == max_date.year:
        filename = f"season_{min_date.year}_accounting_summary.xlsx"
    else:
        filename = f"season_{min_date.year}_{max_date.year}_accounting_summary.xlsx"
    return (input_dir / filename).resolve()


def create_summary_workbook(
    input_files: List[Path],
    income_rows: List[Dict[str, Any]],
    expense_rows: List[Dict[str, Any]],
    daily_rows: List[Dict[str, Any]],
    output_path: Path,
) -> None:
    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    min_date = min(row["date"] for row in daily_rows)
    max_date = max(row["date"] for row in daily_rows)

    total_sales_all = rounded(sum(to_number(r["recorded_total"]) for r in income_rows))
    total_cash_sales = rounded(sum(to_number(r["cash_income"]) for r in daily_rows))
    total_expenses_all = rounded(sum(to_number(r["amount"]) for r in expense_rows))
    total_cash_expenses = rounded(sum(to_number(r["cash_expense"]) for r in daily_rows))
    total_money_to_bank = rounded(sum(to_number(r["money_to_bank"]) for r in daily_rows))
    sales_minus_expenses = rounded(total_sales_all - total_expenses_all)

    # Overview
    ws = wb.create_sheet("Overview")
    style_title(ws, "Accounting Summary from Daily Files", 4)

    overview_rows = [
        ["Metric", "Value"],
        ["Scope", f"{len(input_files)} daily files found in the folder"],
        ["Coverage", f"{min_date.isoformat()} to {max_date.isoformat()}"],
        ["Note", "Based on all matching daily files found in the selected folder."],
        ["Recorded sales (all payment methods)", total_sales_all],
        ["Cash sales / cash income only", total_cash_sales],
        ["Recorded expenses (all payment methods)", total_expenses_all],
        ["Cash expenses only", total_cash_expenses],
        ["Money moved to bank", total_money_to_bank],
        ["Recorded sales minus recorded expenses", sales_minus_expenses],
    ]
    for row in overview_rows:
        ws.append(row)
    style_header_row(ws, 2, 2)
    for r in range(6, 11):
        ws.cell(r, 2).number_format = '#,##0;[Red](#,##0);-'
    ws.cell(11, 2).number_format = '#,##0;[Red](#,##0);-'

    # Period summary section
    start_row = ws.max_row + 3
    ws.cell(start_row, 1, "Summary by Period")
    ws.cell(start_row, 1).font = BOLD_FONT
    ws.cell(start_row, 1).fill = SECTION_FILL
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=4)

    period_header_row = start_row + 1
    for i, header in enumerate(["Period", "Sales (All Methods)", "Expenses (All Methods)", "Sales - Expenses"], start=1):
        ws.cell(period_header_row, i, header)
    style_header_row(ws, period_header_row, 4)

    income_by_date = defaultdict(float)
    for row in income_rows:
        income_by_date[row["date"]] += to_number(row["recorded_total"])

    expense_by_date = defaultdict(float)
    for row in expense_rows:
        expense_by_date[row["date"]] += to_number(row["amount"])

    period_stats = defaultdict(lambda: {"sales": 0.0, "expenses": 0.0})
    for row in daily_rows:
        period = row["period"]
        dt = row["date"]
        period_stats[period]["sales"] += income_by_date.get(dt, 0.0)
        period_stats[period]["expenses"] += expense_by_date.get(dt, 0.0)

    period_items = []
    for period, values in period_stats.items():
        sales = rounded(values["sales"])
        expenses = rounded(values["expenses"])
        period_items.append((period, sales, expenses, rounded(sales - expenses)))
    period_items.sort(key=lambda x: x[0])

    row_idx = period_header_row + 1
    for period, sales, expenses, diff in period_items:
        ws.cell(row_idx, 1, period)
        ws.cell(row_idx, 2, sales)
        ws.cell(row_idx, 3, expenses)
        ws.cell(row_idx, 4, diff)
        for col in (2, 3, 4):
            ws.cell(row_idx, col).number_format = '#,##0;[Red](#,##0);-'
        row_idx += 1

    ws.freeze_panes = "A3"
    autofit_columns(ws)

    # Daily Cash Summary
    daily_table = [
        [
            row["date"],
            row["period"],
            row["opening_balance"],
            row["cash_income"],
            row["cash_expense"],
            row["money_to_bank"],
            row["net_cash_before_bank"],
            row["net_cash_after_bank"],
            row["closing_balance"],
        ]
        for row in sorted(daily_rows, key=lambda x: x["date"])
    ]
    write_table_sheet(
        wb,
        "Daily Cashbox Summary",
        "Daily Cash Summary",
        [
            "Date", "Period", "Opening Balance", "Cash Income", "Cash Expense",
            "Money to Bank", "Net Cash Before Bank", "Net Cash After Bank", "Closing Balance"
        ],
        daily_table,
        date_cols=(1,),
        currency_cols=(3, 4, 5, 6, 7, 8, 9),
    )

    # Sales by Method
    sales_by_method = []
    for method, entries, amount in group_sum_count(income_rows, "payment_method", "recorded_total"):
        share = (amount / total_sales_all) if total_sales_all else 0
        sales_by_method.append([method, entries, amount, share])
    write_table_sheet(
        wb,
        "Recorded Sales by Payment Method",
        "Sales by Method",
        ["Payment Method", "Entries", "Amount", "Share %"],
        sales_by_method,
        currency_cols=(3,),
        percent_cols=(4,),
    )

    # Expenses by Category
    expenses_by_category = [[cat, entries, amount] for cat, entries, amount in group_sum_count(expense_rows, "expense_category", "amount")]
    write_table_sheet(
        wb,
        "Recorded Expenses by Category",
        "Expenses by Category",
        ["Expense Category", "Entries", "Amount"],
        expenses_by_category,
        currency_cols=(3,),
    )

    # Expenses by Method
    expenses_by_method = []
    for method, entries, amount in group_sum_count(expense_rows, "payment_method", "amount"):
        share = (amount / total_expenses_all) if total_expenses_all else 0
        expenses_by_method.append([method, entries, amount, share])
    write_table_sheet(
        wb,
        "Recorded Expenses by Payment Method",
        "Expenses by Method",
        ["Payment Method", "Entries", "Amount", "Share %"],
        expenses_by_method,
        currency_cols=(3,),
        percent_cols=(4,),
    )

    # Sales by Name
    sales_by_name = [[name, entries, amount] for name, entries, amount in group_sum_count(income_rows, "name", "recorded_total")]
    write_table_sheet(
        wb,
        "Top Recorded Sales by Name / Source",
        "Sales by Name",
        ["Name / Source", "Entries", "Amount"],
        sales_by_name,
        currency_cols=(3,),
    )

    # Income Transactions
    income_rows_sorted = sorted(
        income_rows,
        key=lambda x: (
            x["date"],
            (x["name"] or "").lower(),
            (x["payment_method"] or "").lower(),
            to_number(x["recorded_total"]),
        ),
    )
    income_table = [
        [
            row["date"],
            row["name"],
            row["payment_status"],
            row["services"],
            row["payment_method"],
            row["diving_fees"],
            row["extra_misc"],
            row["deposit"],
            row["discount"],
            row["recorded_total"],
        ]
        for row in income_rows_sorted
    ]
    write_table_sheet(
        wb,
        "Income / Sales Transactions",
        "Income Transactions",
        [
            "Date", "Name / Source", "Payment Status", "Services", "Payment Method",
            "Diving Fees", "Extra (Misc)", "Deposit", "Discount", "Recorded Total"
        ],
        income_table,
        date_cols=(1,),
        currency_cols=(6, 7, 8, 9, 10),
    )

    # Expense Transactions
    expense_rows_sorted = sorted(
        expense_rows,
        key=lambda x: (
            x["date"],
            (x["expense_category"] or "").lower(),
            (x["detail"] or "").lower(),
            to_number(x["amount"]),
        ),
    )
    expense_table = [
        [row["date"], row["expense_category"], row["detail"], row["amount"], row["payment_method"]]
        for row in expense_rows_sorted
    ]
    write_table_sheet(
        wb,
        "Expense Transactions",
        "Expense Transactions",
        ["Date", "Expense Category", "Detail", "Amount", "Payment Method"],
        expense_table,
        date_cols=(1,),
        currency_cols=(4,),
    )

    # Save
    wb.save(output_path)


def find_input_files(input_dir: Path) -> List[Path]:
    files: List[Path] = []
    seen = set()
    for pattern in FILE_PATTERNS:
        for path in sorted(input_dir.glob(pattern)):
            if path.name.startswith("~$"):
                continue
            if path.suffix.lower() not in {".xlsm", ".xlsx"}:
                continue
            resolved = path.resolve()
            if resolved not in seen:
                files.append(path)
                seen.add(resolved)
    files.sort(key=parse_file_date)
    return files


def main() -> None:
    parser = argparse.ArgumentParser(description="Build an accounting summary workbook from daily files.")
    parser.add_argument("--input-dir", default=".", help="Folder containing the daily files. Default: current folder")
    parser.add_argument("--output", default=None, help="Output workbook filename. Default: season_<years>_accounting_summary.xlsx")
    args = parser.parse_args()

    input_dir = Path(args.input_dir).resolve()
    if not input_dir.exists() or not input_dir.is_dir():
        raise SystemExit(f"Input directory does not exist or is not a folder: {input_dir}")

    input_files = find_input_files(input_dir)
    if not input_files:
        raise SystemExit(
            "No matching daily files were found.\n"
            "Expected files like: 'Daily Finances 2025-09-03.xlsm'"
        )

    income_rows: List[Dict[str, Any]] = []
    expense_rows: List[Dict[str, Any]] = []
    daily_rows: List[Dict[str, Any]] = []

    for path in input_files:
        wb = load_workbook(path, data_only=True)
        ws = safe_sheet(wb)
        file_date = parse_file_date(path)

        income_rows.extend(extract_income_rows(ws, file_date))
        expense_rows.extend(extract_expense_rows(ws, file_date))
        daily_rows.append(extract_daily_cashbox(ws, file_date))

    if not daily_rows:
        raise SystemExit("No daily cashbox rows could be extracted.")

    assign_periods(daily_rows)

    min_date = min(row["date"] for row in daily_rows)
    max_date = max(row["date"] for row in daily_rows)
    output_path = build_output_path(input_dir, min_date, max_date, args.output)

    create_summary_workbook(
        input_files=input_files,
        income_rows=income_rows,
        expense_rows=expense_rows,
        daily_rows=daily_rows,
        output_path=output_path,
    )

    print(f"Processed {len(input_files)} daily files.")
    print(f"Income transactions: {len(income_rows)}")
    print(f"Expense transactions: {len(expense_rows)}")
    print(f"Summary workbook created: {output_path}")


if __name__ == "__main__":
    main()
